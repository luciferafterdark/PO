import os
from openpyxl import load_workbook

folder_path = "C:/Users/User/OneDrive/Desktop/PO Folders"
excel_path = "C:/Users/User/OneDrive/Desktop/ooredoo/urbanco/PO Sheet.xlsx"

workbook = load_workbook(excel_path)
sheet = workbook.active

invoice_files = os.listdir(folder_path)

for row in sheet.iter_rows(values_only=True):
    po_number = row[5]
    po_status = "INVOICED"

    if f"{po_number}.pdf" in invoice_files:
        po_status = "TO PAY"

    row_list = list(row)
    row_list.append(po_status)
    sheet.append(row_list)

print("done")

workbook.save(excel_path)